from openpyxl import load_workbook
from config import FORMULARIO_ORIGEM, MAPEAMENTO

def ler_dados_excel():
    # Carrega o workbook
    wb = load_workbook(filename=FORMULARIO_ORIGEM, data_only=True)
    ws = wb.active
    
    # Cria um dicionário para armazenar os dados
    dados = {}
    
    # Função auxiliar para obter o valor de uma célula, considerando células mescladas
    def get_cell_value(cell):
        if cell.coordinate in ws.merged_cells:
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    return ws.cell(merged_range.min_row, merged_range.min_col).value
        return cell.value

    # Itera sobre o mapeamento e extrai os valores
    for origem, nome_campo in MAPEAMENTO.items():
        cell = ws[origem]
        valor = get_cell_value(cell)
        
        # Se o valor for None, tenta ler as células à direita
        if valor is None:
            col = cell.column
            while valor is None and col <= ws.max_column:
                valor = get_cell_value(ws.cell(row=cell.row, column=col))
                col += 1
        
        # Se ainda for None, define como uma string vazia
        if valor is None:
            valor = ""
        
        dados[nome_campo] = valor
    
    return dados

def imprimir_dados(dados):
    for campo, valor in dados.items():
        print(f"{campo}: {valor}")

if __name__ == "__main__":
    dados = ler_dados_excel()
    imprimir_dados(dados)