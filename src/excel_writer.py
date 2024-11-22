from openpyxl import load_workbook
from config import FORMULARIO_DESTINO, MAPEAMENTO, FORMULARIO_ORIGEM

# Abrindo a planilha de destino
workbook = load_workbook(FORMULARIO_DESTINO)
sheet = workbook.active

wb = load_workbook(filename=FORMULARIO_ORIGEM, data_only=True)
ws = wb.active
def get_cell_value(cell):
        if cell.coordinate in ws.merged_cells:
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    return ws.cell(merged_range.min_row, merged_range.min_col).value
        return cell.value
# Preenchendo células específicas
#sheet = ORIGEM -> ws = DESTINO
sheet['C10'] = get_cell_value(ws['G8'])
sheet['C13'] = get_cell_value(ws['F10']) + ',' + get_cell_value(ws['AF10']) + ',' + get_cell_value(ws['F11']) + ',' + get_cell_value(ws['AA12'])
sheet['D15'] = get_cell_value(ws['E12'])
sheet['I15'] = get_cell_value(ws['W11'])
sheet['R10'] = get_cell_value(ws['F9'])
sheet['Q15'] = get_cell_value(ws['P12'])
sheet['V15'] = get_cell_value(ws['P4'])
sheet['Z17'] = get_cell_value(ws['Z7'])
sheet['AC9'] = get_cell_value(ws['S9'])
sheet['AC10'] = get_cell_value(ws['AE9'])
sheet['T13'] = get_cell_value(ws['AD4'])
sheet['G25'] = get_cell_value(ws['AA13'])
sheet['P33'] = get_cell_value(ws['AD17'])
sheet['Y33'] = get_cell_value(ws['AD18'])


# sheet['A2'] = 'João'
# sheet['B2'] = 25
# sheet['C2'] = 'São Paulo'

# sheet['A3'] = 'Maria'
# sheet['B3'] = 30
# sheet['C3'] = 'Rio de Janeiro'

# Salvando a planilha
workbook.save('planilha_destino_atualizada.xlsx')